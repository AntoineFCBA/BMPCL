
# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook    
from openpyxl import load_workbook
import xlrd

 ############################## WINDOW ##############################
root = tk.Tk()
root.title('BMPCL')
root.geometry('700x350+20+20')
root.resizable(0, 0)
   

# ================================== directory ==================================
def rep():  
# directory
    print('----------------')
    full_path = filedialog.askdirectory()  # path of the folder
    os.chdir(full_path)
    print("Répertoire actuel : " + full_path)
    print('')
    

def fichier():
    print('----------------')

    # pour insérer le chemin du fichier
    full_path = filedialog.askopenfilename()
    print("Chemin : " + full_path)
    filename = os.path.basename(full_path)
    TB0.insert('end', full_path)
    # -----------------------------------------------

    # read output file avec un délimiteur tabulation
    df = pd.read_csv(full_path, delimiter=",", header=None, engine='python')

    ep_nom = TB3.get(1.0, "end-1c")  # épaisseur nominale
    ep_nom = float(ep_nom)
    larg_nom = TB4.get(1.0, "end-1c")  # largeur nominale
    larg_nom = float(larg_nom)
    
    ep_etalon = TB1.get(1.0, "end-1c")  # épaisseur nominale
    ep_etalon = float(ep_etalon)
    larg_etalon= TB2.get(1.0, "end-1c")  # largeur nominale
    larg_etalon = float(larg_etalon)

    lt = df[0]
    d1 = df[1]
    d2 = df[2]
    d3 = df[3]
    d4 = df[4]
    
    # formule pour calculer l'épaisseur
    epai = ep_etalon - d4 - d2
    # formule pour calculer largeur
    larg = larg_etalon - d1 - d3

    #boucle pour vérifier si l'épaisseur est proche de la valeur nominale
    i=0
    for n in epai:
        if ((ep_nom-2) < epai[i] < (ep_nom + 2)) and ((larg_nom-2) < larg[i] < (larg_nom +2)):
            epai[i] = epai[i]
            larg[i] = larg[i]
           
        else:
            epai[i] = np.nan
            larg[i] = np.nan
            d1[i]   = np.nan
            d2[i]   = np.nan
            d3[i]   = np.nan
            d4[i]   = np.nan
            
    i=i+1
        
    
    m_epai = np.nanmean(epai) if len(epai)>0 else np.nan
    m_larg = np.nanmean(larg) if len(larg)>0 else np.nan
      
    print(larg)
    print(epai)

    print(m_larg,m_epai)
    df.insert(5, "epaisseur", epai, True)
    df.insert(6, "largeur", larg, True)
    df.insert(7, "Epaisseur moyenne", m_epai, True)
    df.insert(8, "Largeur moyenne", m_larg, True)

    df["Epaisseur moyenne"].loc[1:len(df)] = None
    df["Largeur moyenne"].loc[1:len(df)] = None

    df.columns = ['t(s)', 'CH5(mm)', 'CH6(mm)', 'CH7(mm)', 'CH8(mm)', 'Epaisseur(mm)','largeur(mm)', 'Epaisseur moyenne (mm)', 'Largeur moyenne (mm)']
                  
    df.to_csv('num_'+filename, sep=';', decimal=',', header=True, index=False, index_label=None)
             
    print('Terminé')
    
def open_empty_excel():
  workbook = Workbook()
  sheet = workbook.active
  sheet.title = "N° Eprouvette"
        
        # Sauvegarder le classeur Excel vide
  empty_excel_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
  workbook.save(empty_excel_path)
  workbook.close()
        
        # Ouvrir le classeur Excel vide dans l'application par défaut
  os.system(f'start excel.exe "{empty_excel_path}"')
  
def importer():
    import_file_path = filedialog.askopenfile(filetypes=[("Excel Files","*.csv")])
    print(import_file_path)
    if import_file_path:
        workbook = load_workbook(import_file_path)
        print(workbook)
        sheet = workbook.active
        last_row = sheet.max_row
        column_data = [sheet.cell(row=i,column=1).value for i in range(1,last_row+1)]
        csv_file_path = TB0.get(1.0,"end-1c")
        df = pd.read_csv(csv_file_path,delimier = ";",decimal = ",",header = None)
        df.insert(10,"Imported Data",column_data)
      #  df.to_csv('imported_'+os.path.basename(csv_file_path)),sep=';',decimal=',',header = None,index = False)
        df.to_csv('imported_' + os.path.basename(csv_file_path), sep=';', decimal=',', header=None, index=False)
        print("Impration termine")
    
    ############################## LABEL ##############################
L0 = tk.Label(root, text="Chemin du fichier : ").place(x=20, y=20)
L3 = tk.Label(root, text="Epaisseur de l'etalon (mm) : ").place(x=20, y=60)
L4 = tk.Label(root, text="Largeur de l'etalon (mm) : ").place(x=20, y=100)
L1 = tk.Label(root, text="Epaisseur nominale (mm) : ").place(x=20, y=140)
L2 = tk.Label(root, text="Largeur nominale (mm) : ").place(x=20, y=180)

# ================================== choisir ==================================
file_button = tk.Button(root, text='Choisir fichier',bg='red', fg='white', command=fichier)
                
file_button.place(x=70, y=260)
# ================================== repertoire ==================================
file_button = tk.Button(root, text='Choisir Repertoire',bg='green', fg='white', command=rep)
                
file_button.place(x=240, y=260)

#===================================empty xl==========================
file_button = tk.Button(root, text='Ouvrir classeur Excel vide', bg='blue', fg='white', command=open_empty_excel)
file_button.place(x=420, y=260)
#===================================Importer=========================

file_button = tk.Button(root, text = 'Importer', bg= 'grey', fg='white', command = importer)
file_button.place(x=595, y= 260)

############################## TEXT BOX ##############################
TB1 = tk.Text(root, height=1, width=30)
TB1.place(x=190, y=60)

TB2= tk.Text(root, height=1, width=30)
TB2.place(x=190, y=100)

TB3 = tk.Text(root, height=1, width=30)
TB3.place(x=190, y=140)

TB4 = tk.Text(root, height=1, width=30)
TB4.place(x=190, y=180)

TB0 = tk.Text(root, height=1, width=60)
TB0.place(x=190, y=20)

TB3.insert('end', 54)#nominale separeteur de decimale = "."
TB4.insert('end', 154)

TB1.insert('end', 54.15)
TB2.insert('end', 156.90) #etalon separateur de décimale = "."
root.mainloop()
